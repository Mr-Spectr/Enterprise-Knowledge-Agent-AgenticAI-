import unittest

from openpyxl import load_workbook

from response_formatter import create_excel


class ExcelExportTest(unittest.TestCase):
    def test_structured_records_use_a_data_sheet(self):
        path = create_excel("## Student report\nA concise summary.", [{"student_id": "1NT23IS001", "name": "A Vaishnavi", "cgpa": 8.2, "attendance_percent": 96.0, "source": "Moodle SQL import"}])
        try:
            workbook = load_workbook(path)
            self.assertIn("Data", workbook.sheetnames)
            sheet = workbook["Data"]
            self.assertEqual(sheet["A1"].value, "Academic Records - Source-Aware Export")
            self.assertEqual(sheet["A4"].value, "Student ID")
            self.assertEqual(sheet["B5"].value, "A Vaishnavi")
            self.assertEqual(sheet.freeze_panes, "A5")
            self.assertIn("AcademicRecords", sheet.tables)
            self.assertEqual(sheet["G5"].number_format, "0.0%")
            self.assertEqual(sheet.max_column, 15)
        finally:
            path.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
